"""
Export utilities for generating PDF, Excel, and CSV files
"""
import csv
import io
from typing import List, Dict, Any
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd


class BaseExporter:
    """Base class for all exporters"""
    
    def __init__(self, data: List[Dict[str, Any]], filename: str):
        self.data = data
        self.filename = filename
    
    def export(self) -> HttpResponse:
        """Override this method in subclasses"""
        raise NotImplementedError


class PDFExporter(BaseExporter):
    """Export data as PDF using ReportLab"""
    
    def __init__(self, data: List[Dict[str, Any]], filename: str, title: str = "Export", headers: List[str] = None, company_name: str = ""):
        super().__init__(data, filename)
        self.title = title
        self.headers = headers
        self.company_name = company_name
    
    def export(self) -> HttpResponse:
        """Generate PDF using ReportLab"""
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Company name
        if self.company_name:
            company_style = ParagraphStyle(
                'CompanyStyle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=5
            )
            elements.append(Paragraph(self.company_name, company_style))
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#333333'),
            spaceAfter=12,
            alignment=1  # Center
        )
        elements.append(Paragraph(self.title, title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Prepare table data
        if self.data:
            # Get headers from first data item if not provided
            if not self.headers:
                self.headers = list(self.data[0].keys())
            
            # Create table data
            table_data = [self.headers]
            for row in self.data:
                table_data.append([str(row.get(key, '-')) for key in self.headers])
            
            # Create table
            table = Table(table_data)
            
            # Style the table
            table.setStyle(TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Body style
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ]))
            
            elements.append(table)
            
            # Add footer with count
            elements.append(Spacer(1, 0.5*cm))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                alignment=1  # Center
            )
            elements.append(Paragraph(f"Total: {len(self.data)} enregistrement(s)", footer_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF content
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Create HTTP response
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.pdf"'
        
        return response


class ExcelExporter(BaseExporter):
    """Export data as Excel (.xlsx)"""
    
    def __init__(self, data: List[Dict[str, Any]], filename: str, sheet_name: str = 'Data', headers: List[str] = None):
        super().__init__(data, filename)
        self.sheet_name = sheet_name
        self.headers = headers
    
    def export(self) -> HttpResponse:
        """Generate Excel file"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_name
        
        if not self.data:
            return HttpResponse('No data to export', status=400)
        
        # Get headers
        if not self.headers:
            self.headers = list(self.data[0].keys())
        
        # Write headers
        for col_num, header in enumerate(self.headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, row_data in enumerate(self.data, 2):
            for col_num, header in enumerate(self.headers, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = row_data.get(header, '-')
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.xlsx"'
        
        return response


class CSVExporter(BaseExporter):
    """Export data as CSV"""
    
    def __init__(self, data: List[Dict[str, Any]], filename: str, headers: List[str] = None):
        super().__init__(data, filename)
        self.headers = headers
    
    def export(self) -> HttpResponse:
        """Generate CSV file"""
        if not self.data:
            return HttpResponse('No data to export', status=400)
        
        # Get headers
        if not self.headers:
            self.headers = list(self.data[0].keys())
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=self.headers)
        
        writer.writeheader()
        for row in self.data:
            writer.writerow({key: row.get(key, '-') for key in self.headers})
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.csv"'
        
        return response
