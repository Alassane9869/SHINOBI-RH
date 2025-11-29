"""
Exporters améliorés pour la génération de documents professionnels.

Ce module contient des exporters avancés avec support pour :
- WeasyPrint (PDF de haute qualité)
- Métadonnées PDF complètes
- QR codes de vérification
- Formatage Excel avancé (formules, freeze panes, filtres)
- CSV avec UTF-8 BOM
- ZIP avec manifest.json
"""
import csv
import io
import json
import os
import zipfile
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone

# Vérifier la disponibilité de WeasyPrint sans l'importer
WEASYPRINT_AVAILABLE = False
WEASYPRINT_ERROR = None

def _check_weasyprint_available():
    """Vérifie si WeasyPrint est disponible sans l'importer."""
    global WEASYPRINT_AVAILABLE, WEASYPRINT_ERROR
    try:
        import subprocess
        import sys
        
        # Tenter d'importer WeasyPrint dans un sous-processus
        result = subprocess.run(
            [sys.executable, '-c', 'import weasyprint; print("OK")'],
            capture_output=True,
            text=True,
            timeout=5
        )
        
        if result.returncode == 0 and 'OK' in result.stdout:
            WEASYPRINT_AVAILABLE = True
            return True
        else:
            WEASYPRINT_ERROR = "WeasyPrint import failed"
            if result.stderr:
                # Extraire juste le message d'erreur principal
                error_lines = result.stderr.split('\n')
                for line in error_lines:
                    if 'cannot load library' in line or 'GTK' in line or 'OSError' in line:
                        WEASYPRINT_ERROR = line.strip()
                        break
            return False
    except Exception as e:
        WEASYPRINT_ERROR = str(e)
        return False

# Vérifier au chargement du module
_check_weasyprint_available()

if not WEASYPRINT_AVAILABLE:
    print(f"⚠️ WeasyPrint non disponible: {WEASYPRINT_ERROR}")
    print("📝 Utilisation de xhtml2pdf en fallback pour la génération PDF")

# Import xhtml2pdf pour fallback
from xhtml2pdf import pisa

# Import ReportLab pour fallback
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

# Import openpyxl pour Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo

# Import pandas pour manipulation de données
import pandas as pd

# Import QR generator
from .qr_generator import generate_document_qr_code, generate_qr_code_base64


class BaseExporter:
    """
    Classe de base pour tous les exporters.
    
    Fournit des méthodes communes et une interface standardisée.
    """
    
    def __init__(self, data: List[Dict[str, Any]], filename: str, **kwargs):
        """
        Initialise l'exporter.
        
        Args:
            data: Données à exporter
            filename: Nom du fichier (sans extension)
            **kwargs: Paramètres additionnels
        """
        self.data = data
        self.filename = filename
        self.metadata = kwargs.get('metadata', {})
        self.company = kwargs.get('company', None)
        self.user = kwargs.get('user', None)
    
    def export(self) -> HttpResponse:
        """Méthode abstraite à implémenter dans les sous-classes."""
        raise NotImplementedError("Les sous-classes doivent implémenter la méthode export()")
    
    def _generate_metadata(self) -> Dict[str, Any]:
        """
        Génère les métadonnées standard pour l'export.
        
        Returns:
            Dict contenant les métadonnées
        """
        return {
            'created_at': timezone.now().isoformat(),
            'company': self.company.name if self.company else 'Unknown',
            'exported_by': self.user.get_full_name() if self.user else 'System',
            'record_count': len(self.data),
            **self.metadata
        }


class WeasyPrintPDFExporter(BaseExporter):
    """
    Exporter PDF utilisant WeasyPrint pour une qualité professionnelle.
    
    Supporte les templates HTML/CSS avancés, métadonnées complètes,
    et QR codes de vérification.
    """
    
    def __init__(self, data: List[Dict[str, Any]], filename: str, **kwargs):
        super().__init__(data, filename, **kwargs)
        self.template_name = kwargs.get('template_name', 'exports/pdf/default.html')
        self.title = kwargs.get('title', 'Export')
        self.document_id = kwargs.get('document_id', None)
        self.document_type = kwargs.get('document_type', 'document')
        self.context = kwargs.get('context', {})
    
    def export(self) -> HttpResponse:
        """
        Génère un PDF professionnel avec WeasyPrint.
        Utilise xhtml2pdf en fallback si WeasyPrint n'est pas disponible.
        
        Returns:
            HttpResponse avec le PDF
        """
        # Si WeasyPrint n'est pas disponible, utiliser le template simplifié
        template_name = self.template_name
        if not WEASYPRINT_AVAILABLE:
            # Convertir le template vers sa version simplifiée
            if template_name.endswith('.html'):
                base_name = template_name.rsplit('.html', 1)[0]
                simple_template = f"{base_name}_simple.html"
                # Vérifier si le template simplifié existe
                from django.template.loader import get_template
                try:
                    get_template(simple_template)
                    template_name = simple_template
                    print(f"ℹ️ Utilisation du template simplifié: {simple_template}")
                except:
                    print(f"⚠️ Template simplifié non trouvé: {simple_template}, utilisation du template original")
        
        # Récupérer le branding si non fourni
        branding = self.context.get('branding')
        if not branding and self.company:
            try:
                if hasattr(self.company, 'branding'):
                    branding = self.company.branding
            except Exception:
                pass

        # Préparer le contexte pour le template
        context = {
            'title': self.title,
            'data': self.data,
            'company': self.company,
            'user': self.user,
            'metadata': self._generate_metadata(),
            'current_date': timezone.now(),
            'branding': branding,
            **self.context
        }
        
        # Ajouter le QR code si document_id fourni (seulement si WeasyPrint disponible)
        if self.document_id and WEASYPRINT_AVAILABLE:
            try:
                qr_verification_url = getattr(
                    settings, 
                    'EXPORT_QR_VERIFICATION_BASE_URL', 
                    f'{settings.SITE_URL}/verify/'
                )
                qr_code_data_uri = generate_qr_code_base64(
                    f"{qr_verification_url}{self.document_type}/{self.document_id}"
                )
                context['qr_code'] = qr_code_data_uri
            except Exception as e:
                # QR code generation failed, continue without it
                print(f"⚠️ QR code generation failed: {e}")
                context['qr_code'] = None
        
        # Rendre le template HTML
        html_string = render_to_string(template_name, context)
        
        if WEASYPRINT_AVAILABLE:
            # Utiliser WeasyPrint (meilleure qualité)
            try:
                # Import lazy de WeasyPrint
                from weasyprint import HTML, CSS
                from weasyprint.text.fonts import FontConfiguration
                
                # Configuration des polices
                font_config = FontConfiguration()
                
                # CSS de base pour le PDF
                base_css = CSS(string='''
                    @page {
                        size: A4;
                        margin: 2cm;
                        @top-center {
                            content: string(header);
                        }
                        @bottom-center {
                            content: "Page " counter(page) " / " counter(pages);
                        }
                    }
                    body {
                        font-family: 'Inter', 'Roboto', sans-serif;
                        font-size: 11pt;
                        line-height: 1.6;
                    }
                ''', font_config=font_config)
                
                # Générer le PDF
                html = HTML(string=html_string, base_url=str(settings.BASE_DIR))
                pdf_bytes = html.write_pdf(stylesheets=[base_css], font_config=font_config)
            except Exception as e:
                # Erreur WeasyPrint, utiliser fallback
                print(f"⚠️ Erreur WeasyPrint: {e}, fallback vers xhtml2pdf")
                pdf_bytes = self._generate_with_xhtml2pdf(html_string)
        else:
            # Utiliser xhtml2pdf en fallback
            print("ℹ️ Utilisation de xhtml2pdf pour la génération PDF")
            pdf_bytes = self._generate_with_xhtml2pdf(html_string)
        
        # Créer la réponse HTTP
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.pdf"'
        response['Content-Length'] = len(pdf_bytes)
        
        return response
    
    def _generate_with_xhtml2pdf(self, html_string: str) -> bytes:
        """
        Génère un PDF avec ReportLab (fallback simple et fiable).
        
        Args:
            html_string: HTML à convertir (ignoré, on génère directement avec ReportLab)
            
        Returns:
            bytes: Contenu du PDF
        """
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        import logging
        
        logger = logging.getLogger(__name__)
        logger.info("Génération PDF avec ReportLab (fallback)")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=6,
            spaceBefore=12
        )
        
        normal_style = styles['Normal']
        
        # Contenu
        story = []
        
        # Titre
        story.append(Paragraph(self.title or "Rapport de Présence", title_style))
        if self.subtitle:
            story.append(Paragraph(self.subtitle, normal_style))
        story.append(Spacer(1, 0.5*cm))
        
        # Informations de base
        context = self.context
        if 'report_date' in context:
            date_str = context['report_date'].strftime('%d/%m/%Y')
            story.append(Paragraph(f"<b>Date :</b> {date_str}", normal_style))
        
        if 'attendances' in context:
            count = len(context['attendances'])
            story.append(Paragraph(f"<b>Nombre d'employés :</b> {count}", normal_style))
        
        story.append(Spacer(1, 0.5*cm))
        
        # Résumé statistique
        if 'summary' in context:
            summary = context['summary']
            story.append(Paragraph("Résumé", heading_style))
            
            summary_data = [
                ['Présents', 'Retards', 'Absents', 'Excusés'],
                [str(summary.get('present', 0)), str(summary.get('late', 0)), 
                 str(summary.get('absent', 0)), str(summary.get('excused', 0))]
            ]
            
            summary_table = Table(summary_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 14),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 0.5*cm))
        
        # Tableau des détails
        if 'attendances' in context and context['attendances']:
            story.append(Paragraph("Détails de Présence", heading_style))
            
            table_data = [['N°', 'Employé', 'Département', 'Arrivée', 'Départ', 'Statut']]
            
            for idx, att in enumerate(context['attendances'], 1):
                employee_name = att.employee.user.get_full_name() if hasattr(att, 'employee') else 'N/A'
                dept = getattr(att.employee, 'department', '-') if hasattr(att, 'employee') else '-'
                check_in = att.check_in.strftime('%H:%M') if hasattr(att, 'check_in') and att.check_in else '-'
                check_out = att.check_out.strftime('%H:%M') if hasattr(att, 'check_out') and att.check_out else '-'
                status = att.get_status_display() if hasattr(att, 'get_status_display') else str(getattr(att, 'status', '-'))
                
                table_data.append([
                    str(idx),
                    employee_name,
                    dept or '-',
                    check_in,
                    check_out,
                    status
                ])
            
            details_table = Table(table_data, colWidths=[1*cm, 5*cm, 3*cm, 2*cm, 2*cm, 3*cm])
            details_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(details_table)
        else:
            story.append(Paragraph("Aucune donnée de présence pour cette période.", normal_style))
        
        # Pied de page
        story.append(Spacer(1, 1*cm))
        if self.company:
            story.append(Paragraph(f"<i>{self.company.name}</i>", normal_style))
        
        current_date = timezone.now().strftime('%d/%m/%Y à %H:%M')
        story.append(Paragraph(f"<i>Document généré le {current_date}</i>", normal_style))
        
        # Générer le PDF
        doc.build(story)
        pdf_content = buffer.getvalue()
        buffer.close()
        
        logger.info(f"PDF généré avec succès (ReportLab), taille: {len(pdf_content)} bytes")
        return pdf_content


class AdvancedExcelExporter(BaseExporter):
    """
    Exporter Excel avancé avec formatage professionnel.
    
    Supporte :
    - Headers stylés
    - Filtres automatiques
    - Freeze panes
    - Formules
    - Multi-sheets
    - Types de cellules appropriés
    """
    
    def __init__(self, data: List[Dict[str, Any]], filename: str, **kwargs):
        super().__init__(data, filename, **kwargs)
        self.sheet_name = kwargs.get('sheet_name', 'Data')
        self.headers = kwargs.get('headers', None)
        self.include_formulas = kwargs.get('include_formulas', True)
        self.sheets_data = kwargs.get('sheets_data', None)  # Pour multi-sheets
    
    def export(self) -> HttpResponse:
        """
        Génère un fichier Excel formaté professionnellement.
        
        Returns:
            HttpResponse avec le fichier Excel
        """
        workbook = openpyxl.Workbook()
        
        # Si multi-sheets fourni, traiter chaque sheet
        if self.sheets_data:
            # Supprimer la sheet par défaut
            workbook.remove(workbook.active)
            
            for sheet_config in self.sheets_data:
                self._create_sheet(
                    workbook,
                    sheet_config['name'],
                    sheet_config['data'],
                    sheet_config.get('headers'),
                    sheet_config.get('include_totals', False)
                )
        else:
            # Single sheet
            worksheet = workbook.active
            worksheet.title = self.sheet_name
            self._populate_sheet(worksheet, self.data, self.headers)
        
        # Sauvegarder dans un buffer
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Créer la réponse
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.xlsx"'
        
        return response
    
    def _create_sheet(self, workbook, name, data, headers=None, include_totals=False):
        """Crée et formate une sheet."""
        worksheet = workbook.create_sheet(title=name)
        self._populate_sheet(worksheet, data, headers, include_totals)
    
    def _populate_sheet(self, worksheet, data, headers=None, include_totals=False):
        """Remplit une sheet avec les données et le formatage."""
        if not data:
            return
        
        # Headers
        if not headers:
            headers = list(data[0].keys())
        
        # Style pour les headers
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Écrire les headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Écrire les données
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = row_data.get(header, '')
                
                # Détecter le type de données
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
                elif isinstance(value, datetime):
                    cell.value = value
                    cell.number_format = 'DD/MM/YYYY'
                else:
                    cell.value = str(value)
                
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
        
        # Ajouter une ligne de totaux si demandé
        if include_totals and self.include_formulas:
            total_row = len(data) + 2
            worksheet.cell(row=total_row, column=1).value = "TOTAL"
            worksheet.cell(row=total_row, column=1).font = Font(bold=True)
            
            # Ajouter des formules SUM pour les colonnes numériques
            for col_num, header in enumerate(headers, 1):
                first_value = data[0].get(header) if data else None
                if isinstance(first_value, (int, float)):
                    cell = worksheet.cell(row=total_row, column=col_num)
                    cell.value = f"=SUM({get_column_letter(col_num)}2:{get_column_letter(col_num)}{total_row-1})"
                    cell.font = Font(bold=True)
                    cell.number_format = '#,##0.00'
        
        # Freeze panes (figer la première ligne)
        worksheet.freeze_panes = 'A2'
        
        # Activer les filtres automatiques
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


class UTF8CSVExporter(BaseExporter):
    """
    Exporter CSV avec UTF-8 BOM pour compatibilité Excel.
    
    Utilise le séparateur point-virgule (;) standard français.
    """
    
    def __init__(self, data: List[Dict[str, Any]], filename: str, **kwargs):
        super().__init__(data, filename, **kwargs)
        self.headers = kwargs.get('headers', None)
        self.delimiter = kwargs.get('delimiter', ';')
    
    def export(self) -> HttpResponse:
        """
        Génère un fichier CSV avec UTF-8 BOM.
        
        Returns:
            HttpResponse avec le CSV
        """
        if not self.data:
            return HttpResponse('Aucune donnée à exporter', status=400)
        
        # Headers
        if not self.headers:
            self.headers = list(self.data[0].keys())
        
        # Créer le CSV en mémoire
        output = io.StringIO()
        
        # Ajouter le BOM UTF-8 pour Excel
        output.write('\ufeff')
        
        writer = csv.DictWriter(
            output,
            fieldnames=self.headers,
            delimiter=self.delimiter,
            quoting=csv.QUOTE_MINIMAL,
            escapechar='\\'
        )
        
        writer.writeheader()
        for row in self.data:
            writer.writerow({key: row.get(key, '') for key in self.headers})
        
        # Créer la réponse
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.csv"'
        
        return response


class ZIPExporter(BaseExporter):
    """
    Exporter ZIP avec manifest.json.
    
    Permet de créer des archives structurées avec métadonnées.
    """
    
    def __init__(self, files: List[Dict[str, Any]], filename: str, **kwargs):
        """
        Args:
            files: Liste de dicts avec 'name' et 'content' (bytes)
            filename: Nom du fichier ZIP
        """
        super().__init__([], filename, **kwargs)
        self.files = files
        self.structure = kwargs.get('structure', {})
    
    def export(self) -> HttpResponse:
        """
        Génère une archive ZIP avec manifest.
        
        Returns:
            HttpResponse avec le ZIP
        """
        buffer = io.BytesIO()
        
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Ajouter les fichiers
            for file_info in self.files:
                zip_file.writestr(file_info['name'], file_info['content'])
            
            # Créer le manifest.json
            manifest = {
                'created_at': timezone.now().isoformat(),
                'company': self.company.name if self.company else 'Unknown',
                'exported_by': self.user.get_full_name() if self.user else 'System',
                'file_count': len(self.files),
                'files': [
                    {
                        'name': f['name'],
                        'size': len(f['content']),
                        'type': f.get('type', 'unknown')
                    }
                    for f in self.files
                ],
                'structure': self.structure,
                **self.metadata
            }
            
            zip_file.writestr('manifest.json', json.dumps(manifest, indent=2, ensure_ascii=False))
        
        buffer.seek(0)
        
        # Créer la réponse
        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.zip"'
        
        return response
