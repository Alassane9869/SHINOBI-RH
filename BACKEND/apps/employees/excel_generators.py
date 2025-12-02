"""
Générateur Excel pour les listes et rapports d'employés.
"""
import io
from datetime import datetime, date
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule


class EmployeeExcelGenerator:
    def __init__(self, company):
        self.company = company
        
        # Palette de couleurs épurée
        self.color_primary = '2563EB'
        self.color_success = '10B981'
        self.color_warning = 'F59E0B'
        self.color_danger = 'EF4444'
        self.color_text = '1F2937'
        self.color_border = 'E5E7EB'
        self.color_bg_light = 'F9FAFB'
        
        # Styles
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color=self.color_primary, end_color=self.color_primary, fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.border_thin = Border(
            left=Side(style='thin', color=self.color_border),
            right=Side(style='thin', color=self.color_border),
            top=Side(style='thin', color=self.color_border),
            bottom=Side(style='thin', color=self.color_border)
        )

    def _create_workbook(self, title):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employés"
        
        # Titre
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(size=16, bold=True, color=self.color_primary)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Info entreprise
        ws.merge_cells('A2:H2')
        cell = ws['A2']
        cell.value = self.company.name
        cell.font = Font(size=11, color='666666')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[3].height = 5
        
        return wb, ws

    def _style_header(self, cell):
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_alignment
        cell.border = self.border_thin

    def _style_cell(self, cell, center=True, bg_color=None):
        if center:
            cell.alignment = self.center_alignment
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.border_thin
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    def _auto_adjust_columns(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_employee_list(self, data, filename):
        """
        Génère une liste complète des employés.
        
        Args:
            data: {
                'employees': [
                    {
                        'name': 'Jean Dupont',
                        'matricule': 'EMP001',
                        'department': 'Développement',
                        'position': 'Développeur Senior',
                        'hire_date': '01/01/2020',
                        'years_service': 4,
                        'base_salary': 500000,
                        'status': 'Actif'
                    },
                    ...
                ]
            }
        """
        wb, ws = self._create_workbook("Liste des Employés")
        
        current_row = 4
        
        # === 1. KPIs ===
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "📊 VUE D'ENSEMBLE"
        cell.font = Font(size=12, bold=True, color=self.color_primary)
        current_row += 1
        
        employees = data.get('employees', [])
        total_employees = len(employees)
        active_employees = sum(1 for e in employees if e.get('status') == 'Actif')
        avg_years = sum(e.get('years_service', 0) for e in employees) / total_employees if total_employees > 0 else 0
        
        kpis = [
            ('Total Employés', str(total_employees), self.color_primary),
            ('Actifs', str(active_employees), self.color_success),
            ('Ancienneté Moyenne', f"{avg_years:.1f} ans", self.color_primary),
        ]
        
        col = 1
        for label, value, color in kpis:
            cell = ws.cell(row=current_row, column=col)
            cell.value = label
            cell.font = Font(size=9, color='666666')
            cell.alignment = self.center_alignment
            
            cell = ws.cell(row=current_row + 1, column=col)
            cell.value = value
            cell.font = Font(size=16, bold=True, color=color)
            cell.alignment = self.center_alignment
            
            col += 2
        
        current_row += 4
        
        # === 2. Tableau ===
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "📋 LISTE COMPLÈTE"
        cell.font = Font(size=12, bold=True, color=self.color_primary)
        current_row += 1
        
        headers = ['Nom', 'Matricule', 'Département', 'Poste', 'Embauche', 'Ancienneté', 'Salaire', 'Statut']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            self._style_header(cell)
        
        header_row = current_row
        current_row += 1
        
        for emp in employees:
            row_data = [
                emp.get('name', ''),
                emp.get('matricule', ''),
                emp.get('department', '-'),
                emp.get('position', ''),
                emp.get('hire_date', ''),
                f"{emp.get('years_service', 0)} ans",
                emp.get('base_salary', 0),
                emp.get('status', 'Actif')
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = val
                
                # Formatage numérique
                if col == 7 and isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
                
                # Style
                if col <= 4:
                    self._style_cell(cell, center=False)
                else:
                    self._style_cell(cell, center=True)
                
                # Couleur pour statut
                if col == 8:
                    if val == 'Actif':
                        cell.font = Font(bold=True, color=self.color_success)
                    else:
                        cell.font = Font(bold=True, color=self.color_danger)
                
                # Lignes alternées
                if current_row % 2 == 0:
                    self._style_cell(cell, center=(col > 4), bg_color=self.color_bg_light)
            
            current_row += 1
        
        # === 3. Formatage Conditionnel ===
        if current_row > header_row + 1:
            # Barres de données pour salaire
            salary_col = get_column_letter(7)
            ws.conditional_formatting.add(
                f'{salary_col}{header_row + 1}:{salary_col}{current_row - 1}',
                DataBarRule(
                    start_type='min',
                    end_type='max',
                    color=self.color_primary,
                    showValue=True
                )
            )
        
        # === 4. Figer et Filtres ===
        ws.freeze_panes = ws[f'A{header_row + 1}']
        ws.auto_filter.ref = f'A{header_row}:H{current_row - 1}'
        self._auto_adjust_columns(ws)
        
        # === 5. Note ===
        current_row += 2
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        cell.font = Font(size=9, italic=True, color='666666')
        cell.alignment = Alignment(horizontal='center')
        
        return self._save_response(wb, filename)

    def _save_response(self, wb, filename):
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response
