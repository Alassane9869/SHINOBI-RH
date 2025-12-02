import io
from datetime import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule

class AttendanceExcelGenerator:
    def __init__(self, company):
        self.company = company
        
        # Palette de couleurs épurée (minimaliste)
        self.color_primary = '2563EB'      # Bleu
        self.color_success = '10B981'      # Vert
        self.color_warning = 'F59E0B'      # Orange
        self.color_danger = 'EF4444'       # Rouge
        self.color_text = '1F2937'         # Gris foncé
        self.color_border = 'E5E7EB'       # Bordure
        self.color_bg_light = 'F9FAFB'     # Fond clair
        
        # Styles réutilisables
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color=self.color_primary, end_color=self.color_primary, fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.border_thin = Border(
            left=Side(style='thin', color=self.color_border), 
            right=Side(style='thin', color=self.color_border), 
            top=Side(style='thin', color=self.color_border), 
            bottom=Side(style='thin', color=self.color_border)
        )

    def _create_workbook(self, title):
        """Crée un classeur avec en-tête épuré."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"
        
        # Titre principal (fusionné, épuré)
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(size=16, bold=True, color=self.color_primary)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Info entreprise (fusionné, discret)
        ws.merge_cells('A2:G2')
        cell = ws['A2']
        cell.value = self.company.name
        cell.font = Font(size=11, color='666666')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ligne vide pour respiration
        ws.row_dimensions[3].height = 5
        
        return wb, ws

    def _auto_adjust_columns(self, ws):
        """Ajuste automatiquement la largeur des colonnes."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)  # Limite max
            ws.column_dimensions[column_letter].width = adjusted_width

    def _style_header(self, cell):
        """Style pour en-tête de tableau."""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_alignment
        cell.border = self.border_thin

    def _style_cell(self, cell, center=True, bg_color=None):
        """Style pour cellule normale."""
        if center:
            cell.alignment = self.center_alignment
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.border_thin
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    def _create_kpi_section(self, ws, start_row, summary):
        """Crée une section KPI visuelle."""
        # Titre section
        ws.merge_cells(f'A{start_row}:G{start_row}')
        cell = ws[f'A{start_row}']
        cell.value = "📊 VUE D'ENSEMBLE"
        cell.font = Font(size=12, bold=True, color=self.color_primary)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        start_row += 1
        
        # KPIs en cartes (4 colonnes)
        kpis = [
            ('Présents', summary.get('present', 0), self.color_success),
            ('Retards', summary.get('late', 0), self.color_warning),
            ('Absents', summary.get('absent', 0), self.color_danger),
            ('Excusés', summary.get('excused', 0), '6B7280'),  # Gris
        ]
        
        col = 1
        for label, value, color in kpis:
            # Label
            cell = ws.cell(row=start_row, column=col)
            cell.value = label
            cell.font = Font(size=9, color='666666')
            cell.alignment = self.center_alignment
            
            # Valeur (grande et colorée)
            cell = ws.cell(row=start_row + 1, column=col)
            cell.value = value
            cell.font = Font(size=18, bold=True, color=color)
            cell.alignment = self.center_alignment
            
            # Bordure de carte
            for r in range(start_row, start_row + 2):
                ws.cell(row=r, column=col).border = Border(
                    left=Side(style='medium', color=color),
                    right=Side(style='thin', color=self.color_border),
                    top=Side(style='thin', color=self.color_border),
                    bottom=Side(style='thin', color=self.color_border)
                )
            
            col += 2  # Espacement entre cartes
        
        return start_row + 3  # Retourner la prochaine ligne disponible

    def generate_daily_report(self, data, filename):
        """Génère un rapport journalier Excel épuré et professionnel."""
        wb, ws = self._create_workbook(f"Rapport Journalier - {data.get('date')}")
        
        current_row = 4
        
        # === 1. Section KPI ===
        current_row = self._create_kpi_section(ws, current_row, data.get('summary', {}))
        current_row += 1
        
        # === 2. Tableau Détaillé ===
        # Titre section
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "📋 DÉTAIL DES PRÉSENCES"
        cell.font = Font(size=12, bold=True, color=self.color_primary)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        
        # En-têtes
        headers = ['Employé', 'Département', 'Arrivée', 'Départ', 'Statut', 'Retard (min)', 'Heures']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            self._style_header(cell)
        
        header_row = current_row
        current_row += 1
        
        # Données
        for att in data.get('attendances', []):
            row_data = [
                att['Employé'],
                att['Département'],
                att['Arrivée'],
                att['Départ'],
                att['Statut'],
                att['Retard (min)'],
                att['Heures']
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = val
                
                # Style selon la colonne
                if col <= 2:  # Nom et Département
                    self._style_cell(cell, center=False)
                else:
                    self._style_cell(cell, center=True)
                
                # Couleur pour statut
                if col == 5:  # Colonne Statut
                    if 'Absent' in str(val):
                        cell.font = Font(bold=True, color=self.color_danger)
                    elif 'Retard' in str(val) or 'Late' in str(val):
                        cell.font = Font(bold=True, color=self.color_warning)
                    elif 'Présent' in str(val) or 'Present' in str(val):
                        cell.font = Font(bold=True, color=self.color_success)
                
                # Lignes alternées (zebra)
                if current_row % 2 == 0:
                    if not cell.fill or cell.fill.start_color.rgb != self.header_fill.start_color.rgb:
                        self._style_cell(cell, center=(col > 2), bg_color=self.color_bg_light)
            
            current_row += 1
        
        # === 3. Formatage Conditionnel ===
        # Barres de données pour les heures
        if current_row > header_row + 1:
            hours_col = get_column_letter(7)
            ws.conditional_formatting.add(
                f'{hours_col}{header_row + 1}:{hours_col}{current_row - 1}',
                DataBarRule(
                    start_type='min',
                    end_type='max',
                    color=self.color_success,
                    showValue=True
                )
            )
        
        # === 4. Figer les volets ===
        ws.freeze_panes = ws[f'A{header_row + 1}']
        
        # === 5. Filtres automatiques ===
        ws.auto_filter.ref = f'A{header_row}:G{current_row - 1}'
        
        # === 6. Ajuster les colonnes ===
        self._auto_adjust_columns(ws)
        
        # === 7. Note de bas de page ===
        current_row += 2
        total = sum([data['summary'].get(k, 0) for k in ['present', 'late', 'absent', 'excused']])
        if total > 0:
            presence_rate = ((data['summary'].get('present', 0) + data['summary'].get('late', 0)) / total) * 100
            ws.merge_cells(f'A{current_row}:G{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = f"Taux de présence global : {presence_rate:.1f}% • Total employés : {total}"
            cell.font = Font(size=9, italic=True, color='666666')
            cell.alignment = Alignment(horizontal='center')
        
        return self._save_response(wb, filename)

    def generate_monthly_advanced_report(self, data, filename):
        """Génère un rapport mensuel Excel épuré."""
        wb, ws = self._create_workbook(f"Rapport Mensuel - {data.get('month')}")
        
        current_row = 4
        
        # === 1. KPIs Globaux ===
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "📊 PERFORMANCE GLOBALE"
        cell.font = Font(size=12, bold=True, color=self.color_primary)
        current_row += 1
        
        stats = data.get('stats', {})
        kpis = [
            ('Taux de Présence', f"{stats.get('present_rate', 0):.1f}%", self.color_success),
            ('Taux de Retard', f"{stats.get('late_rate', 0):.1f}%", self.color_warning),
            ('Taux d\'Absence', f"{stats.get('absent_rate', 0):.1f}%", self.color_danger),
        ]
        
        col = 1
        for label, value, color in kpis:
            cell = ws.cell(row=current_row, column=col)
            cell.value = label
            cell.font = Font(size=9, color='666666')
            cell.alignment = self.center_alignment
            
            cell = ws.cell(row=current_row + 1, column=col)
            cell.value = value
            cell.font = Font(size=18, bold=True, color=color)
            cell.alignment = self.center_alignment
            
            col += 2
        
        current_row += 4
        
        # === 2. Tableau par Employé ===
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "📋 PERFORMANCE PAR EMPLOYÉ"
        cell.font = Font(size=12, bold=True, color=self.color_primary)
        current_row += 1
        
        headers = ['Employé', 'Département', 'Présent', 'Retard', 'Absent', 'Taux']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            self._style_header(cell)
        
        header_row = current_row
        current_row += 1
        
        for emp in data.get('employee_stats', []):
            row_data = [
                emp['employee_name'],
                emp['department'] or '-',
                emp['present'],
                emp['late'],
                emp['absent'],
                f"{emp['attendance_rate']:.1f}%"
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = val
                self._style_cell(cell, center=(col > 2))
                
                # Couleur pour le taux
                if col == 6:
                    rate = emp['attendance_rate']
                    if rate >= 95:
                        cell.font = Font(bold=True, color=self.color_success)
                    elif rate >= 80:
                        cell.font = Font(bold=True, color=self.color_warning)
                    else:
                        cell.font = Font(bold=True, color=self.color_danger)
                
                if current_row % 2 == 0:
                    self._style_cell(cell, center=(col > 2), bg_color=self.color_bg_light)
            
            current_row += 1
        
        # === 3. Formatage Conditionnel (Échelle de couleurs pour taux) ===
        if current_row > header_row + 1:
            rate_col = get_column_letter(6)
            ws.conditional_formatting.add(
                f'{rate_col}{header_row + 1}:{rate_col}{current_row - 1}',
                ColorScaleRule(
                    start_type='num', start_value=70, start_color='EF4444',  # Rouge
                    mid_type='num', mid_value=85, mid_color='F59E0B',        # Orange
                    end_type='num', end_value=100, end_color='10B981'        # Vert
                )
            )
        
        ws.freeze_panes = ws[f'A{header_row + 1}']
        ws.auto_filter.ref = f'A{header_row}:F{current_row - 1}'
        self._auto_adjust_columns(ws)
        
        return self._save_response(wb, filename)

    def _save_response(self, wb, filename):
        """Sauvegarde et retourne le fichier Excel."""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response
