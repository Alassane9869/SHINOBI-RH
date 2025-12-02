"""
Générateur Excel pour les rapports de congés.
"""
import io
from datetime import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


class LeaveExcelGenerator:
    def __init__(self, company):
        self.company = company
        
        # Palette de couleurs épurée
        self.color_primary = '2563EB'
        self.color_success = '10B981'
        self.color_warning = 'F59E0B'
        self.color_danger = 'EF4444'
        self.color_text = '1F2937'
        self.color_border = 'E5E7EB'
        self.color_bg_light = 'F9FAFB'
        
        # Styles
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color=self.color_primary, end_color=self.color_primary, fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.border_thin = Border(
            left=Side(style='thin', color=self.color_border),
            right=Side(style='thin', color=self.color_border),
            top=Side(style='thin', color=self.color_border),
            bottom=Side(style='thin', color=self.color_border)
        )

    def _create_workbook(self, title):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Congés"
        
        # Titre
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(size=16, bold=True, color=self.color_primary)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Info entreprise
        ws.merge_cells('A2:G2')
        cell = ws['A2']
        cell.value = self.company.name
        cell.font = Font(size=11, color='666666')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[3].height = 5
        
        return wb, ws

    def _style_header(self, cell):
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_alignment
        cell.border = self.border_thin

    def _style_cell(self, cell, center=True, bg_color=None):
        if center:
            cell.alignment = self.center_alignment
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.border_thin
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    def _auto_adjust_columns(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_leave_report(self, data, filename):
        """
        Génère un rapport de congés.
        
        Args:
            data: {
                'period': 'Décembre 2024',
                'leaves': [
                    {
                        'employee_name': 'Jean Dupont',
                        'department': 'Développement',
                        'leave_type': 'Vacation',
                        'start_date': '01/12/2024',
                        'end_date': '15/12/2024',
                        'days': 15,
                        'status': 'Approved'
                    },
                    ...
                ]
            }
        """
        wb, ws = self._create_workbook(f"Rapport de Congés - {data.get('period')}")
        
        current_row = 4
        
        # === 1. KPIs ===
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "📊 VUE D'ENSEMBLE"
        cell.font = Font(size=12, bold=True, color=self.color_primary)
        current_row += 1
        
        leaves = data.get('leaves', [])
        total_leaves = len(leaves)
        approved = sum(1 for l in leaves if l.get('status') == 'Approved')
        pending = sum(1 for l in leaves if l.get('status') == 'Pending')
        rejected = sum(1 for l in leaves if l.get('status') == 'Rejected')
        total_days = sum(l.get('days', 0) for l in leaves if l.get('status') == 'Approved')
        
        kpis = [
            ('Total Demandes', str(total_leaves), self.color_primary),
            ('Approuvées', str(approved), self.color_success),
            ('En attente', str(pending), self.color_warning),
            ('Jours Approuvés', f"{total_days} jours", self.color_success),
        ]
        
        col = 1
        for label, value, color in kpis[:4]:
            cell = ws.cell(row=current_row, column=col)
            cell.value = label
            cell.font = Font(size=9, color='666666')
            cell.alignment = self.center_alignment
            
            cell = ws.cell(row=current_row + 1, column=col)
            cell.value = value
            cell.font = Font(size=14, bold=True, color=color)
            cell.alignment = self.center_alignment
            
            col += 2 if col < 7 else 1
        
        current_row += 4
        
        # === 2. Tableau ===
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "📋 DÉTAIL DES CONGÉS"
        cell.font = Font(size=12, bold=True, color=self.color_primary)
        current_row += 1
        
        headers = ['Employé', 'Département', 'Type', 'Début', 'Fin', 'Jours', 'Statut']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            self._style_header(cell)
        
        header_row = current_row
        current_row += 1
        
        for leave in leaves:
            row_data = [
                leave.get('employee_name', ''),
                leave.get('department', '-'),
                leave.get('leave_type', ''),
                leave.get('start_date', ''),
                leave.get('end_date', ''),
                leave.get('days', 0),
                leave.get('status', '')
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = val
                
                # Style
                if col <= 3:
                    self._style_cell(cell, center=False)
                else:
                    self._style_cell(cell, center=True)
                
                # Couleur pour statut
                if col == 7:
                    if val == 'Approved':
                        cell.font = Font(bold=True, color=self.color_success)
                    elif val == 'Pending':
                        cell.font = Font(bold=True, color=self.color_warning)
                    elif val == 'Rejected':
                        cell.font = Font(bold=True, color=self.color_danger)
                
                # Lignes alternées
                if current_row % 2 == 0:
                    self._style_cell(cell, center=(col > 3), bg_color=self.color_bg_light)
            
            current_row += 1
        
        # === 3. Figer et Filtres ===
        ws.freeze_panes = ws[f'A{header_row + 1}']
        ws.auto_filter.ref = f'A{header_row}:G{current_row - 1}'
        self._auto_adjust_columns(ws)
        
        # === 4. Note ===
        current_row += 2
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Total jours de congés approuvés : {total_days}"
        cell.font = Font(size=9, italic=True, color='666666')
        cell.alignment = Alignment(horizontal='center')
        
        return self._save_response(wb, filename)

    def _save_response(self, wb, filename):
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response
