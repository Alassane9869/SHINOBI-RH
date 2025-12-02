"""
Générateur Excel pour les rapports Owner/Super Admin.
"""
import io
from datetime import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference


class OwnerExcelGenerator:
    def __init__(self, company=None):
        self.company = company
        
        # Palette de couleurs épurée
        self.color_primary = '2563EB'
        self.color_success = '10B981'
        self.color_warning = 'F59E0B'
        self.color_danger = 'EF4444'
        self.color_text = '1F2937'
        self.color_border = 'E5E7EB'
        self.color_bg_light = 'F9FAFB'
        
        # Styles
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color=self.color_primary, end_color=self.color_primary, fill_type='solid')
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.border_thin = Border(
            left=Side(style='thin', color=self.color_border),
            right=Side(style='thin', color=self.color_border),
            top=Side(style='thin', color=self.color_border),
            bottom=Side(style='thin', color=self.color_border)
        )

    def _create_workbook(self, title):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        
        # Titre
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(size=18, bold=True, color=self.color_primary)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Info
        ws.merge_cells('A2:H2')
        cell = ws['A2']
        cell.value = self.company.name if self.company else "Rapport Global"
        cell.font = Font(size=11, color='666666')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[3].height = 5
        
        return wb, ws

    def _style_header(self, cell):
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_alignment
        cell.border = self.border_thin

    def _style_cell(self, cell, center=True, bg_color=None):
        if center:
            cell.alignment = self.center_alignment
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.border_thin
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    def _auto_adjust_columns(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_executive_dashboard(self, data, filename):
        """
        Génère un dashboard exécutif complet.
        
        Args:
            data: {
                'period': 'Novembre 2024',
                'departments': [
                    {
                        'name': 'Développement',
                        'employees': 50,
                        'attendance_rate': 95.2,
                        'payroll': 25000000,
                        'avg_salary': 500000
                    },
                    ...
                ]
            }
        """
        wb, ws = self._create_workbook(f"Dashboard Exécutif - {data.get('period')}")
        
        current_row = 4
        
        # === 1. KPIs Globaux ===
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "📊 VUE D'ENSEMBLE GLOBALE"
        cell.font = Font(size=14, bold=True, color=self.color_primary)
        current_row += 1
        
        departments = data.get('departments', [])
        total_employees = sum(d.get('employees', 0) for d in departments)
        avg_attendance = sum(d.get('attendance_rate', 0) for d in departments) / len(departments) if departments else 0
        total_payroll = sum(d.get('payroll', 0) for d in departments)
        
        kpis = [
            ('Total Employés', str(total_employees), self.color_primary),
            ('Taux Présence Moyen', f"{avg_attendance:.1f}%", self.color_success),
            ('Masse Salariale', f"{total_payroll:,.0f} FCFA", self.color_primary),
        ]
        
        col = 1
        for label, value, color in kpis:
            cell = ws.cell(row=current_row, column=col)
            cell.value = label
            cell.font = Font(size=9, color='666666')
            cell.alignment = self.center_alignment
            
            cell = ws.cell(row=current_row + 1, column=col)
            cell.value = value
            cell.font = Font(size=18, bold=True, color=color)
            cell.alignment = self.center_alignment
            
            col += 3
        
        current_row += 4
        
        # === 2. Tableau par Département ===
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "📋 PERFORMANCE PAR DÉPARTEMENT"
        cell.font = Font(size=12, bold=True, color=self.color_primary)
        current_row += 1
        
        headers = ['Département', 'Employés', 'Taux Présence', 'Masse Salariale', 'Salaire Moyen']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            self._style_header(cell)
        
        header_row = current_row
        current_row += 1
        
        for dept in departments:
            row_data = [
                dept.get('name', ''),
                dept.get('employees', 0),
                dept.get('attendance_rate', 0),
                dept.get('payroll', 0),
                dept.get('avg_salary', 0)
            ]
            
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = val
                
                # Formatage
                if col == 1:
                    self._style_cell(cell, center=False)
                elif col == 3:
                    cell.value = f"{val:.1f}%"
                    self._style_cell(cell, center=True)
                elif col >= 4:
                    cell.number_format = '#,##0'
                    self._style_cell(cell, center=True)
                else:
                    self._style_cell(cell, center=True)
                
                # Lignes alternées
                if current_row % 2 == 0:
                    self._style_cell(cell, center=(col > 1), bg_color=self.color_bg_light)
            
            current_row += 1
        
        # === 3. Ligne de Total ===
        total_row_data = [
            'TOTAL',
            total_employees,
            avg_attendance,
            total_payroll,
            total_payroll / total_employees if total_employees > 0 else 0
        ]
        
        for col, val in enumerate(total_row_data, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = val
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color=self.color_bg_light, end_color=self.color_bg_light, fill_type='solid')
            cell.border = Border(
                top=Side(style='medium', color=self.color_primary),
                bottom=Side(style='medium', color=self.color_primary),
                left=Side(style='thin', color=self.color_border),
                right=Side(style='thin', color=self.color_border)
            )
            
            if col == 3:
                cell.value = f"{val:.1f}%"
                cell.alignment = self.center_alignment
            elif col >= 4:
                cell.number_format = '#,##0'
                cell.alignment = self.center_alignment
            elif col == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # === 4. Formatage Conditionnel ===
        if current_row > header_row + 1:
            # Échelle de couleurs pour taux de présence
            rate_col = get_column_letter(3)
            ws.conditional_formatting.add(
                f'{rate_col}{header_row + 1}:{rate_col}{current_row - 1}',
                ColorScaleRule(
                    start_type='num', start_value=80, start_color='EF4444',
                    mid_type='num', mid_value=90, mid_color='F59E0B',
                    end_type='num', end_value=100, end_color='10B981'
                )
            )
            
            # Barres de données pour masse salariale
            payroll_col = get_column_letter(4)
            ws.conditional_formatting.add(
                f'{payroll_col}{header_row + 1}:{payroll_col}{current_row - 1}',
                DataBarRule(
                    start_type='min',
                    end_type='max',
                    color=self.color_primary,
                    showValue=True
                )
            )
        
        # === 5. Figer et Filtres ===
        ws.freeze_panes = ws[f'A{header_row + 1}']
        ws.auto_filter.ref = f'A{header_row}:E{current_row - 1}'
        self._auto_adjust_columns(ws)
        
        # === 6. Note ===
        current_row += 2
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        cell.font = Font(size=9, italic=True, color='666666')
        cell.alignment = Alignment(horizontal='center')
        
        return self._save_response(wb, filename)

    def _save_response(self, wb, filename):
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response
